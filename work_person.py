import openpyxl
import os

from wordcloud import WordCloud, STOPWORDS
     

import nltk
import re
from nltk.stem.snowball import SnowballStemmer
from nltk.stem import WordNetLemmatizer
from nltk.corpus import stopwords,wordnet

import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import urllib.request
import time

wb= openpyxl.load_workbook(r"C:\Users\신충현\Desktop\project\All_Task.xlsx")
sheet1 = wb.get_sheet_by_name("all_data")

class Kind_person:
    def __init__(self):
            p_who = []       
            for i in list(range(2,sheet1.max_row-1)):
                p_who.append(sheet1.cell(i,2).value)          
            str_who = ",".join(p_who)
            list_who = str_who.split(",")
            self.set_who = set(list_who)

class Person:
    def __init__(self,name): 
        self.name = name
        name_category = {}
        for i in list(range(2,sheet1.max_row-1)):
            if name in sheet1.cell(i,2).value:
                if  sheet1.cell(i,4).value not in name_category:
                     name_category[sheet1.cell(i,4).value] = float(sheet1.cell(i,5).value)
                else:
                    name_category[sheet1.cell(i,4).value] += float(sheet1.cell(i,5).value)
            else:
                continue
        
        #총 분야의 소요시간 합계
        sum_value = 0
        for i in name_category.values():
            sum_value += i
        
        #시간을 percentage 단위로 변환
        for key,element in name_category.items():
            name_category[key] = round((element/sum_value)*100,2)

        #시간에 대해 최댓값을 갖는 3가지 분야만 도출
        final_category = {}
        for i in range(3):
            value_list = list(name_category.values())
            key_list = list(name_category.keys())
            max_key = key_list[value_list.index(max(value_list))]
            final_category[max_key] = max(value_list)
            del name_category[max_key]

        #남은 부분 계산
        three_sum = 0
        for i in final_category.values():
            three_sum += i
        residual = 100 - three_sum

        #남은 부분 추가
        final_category["residual"] = round(residual,2)        
        self.final_category =final_category

    def ratio_Cat(self):
        #입력된 사람의 분야별 소요시간(최대 3)        
        return ",".join("{}:{}".format(key,value) for key, value in self.final_category.items())

    def chart_pie(self):
        labels = tuple(self.final_category.keys())
        sizes = list(self.final_category.values())
        explode = (0, 0.1, 0, 0)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
                shadow=True, startangle=90)
        ax1.axis('equal') 
        
        plt.savefig('{}_04.png'.format(self.name),dpi=300)
        a = '{}_04.png'.format(self.name)
        return a    

    def WordCloud(self):
        max_category = list(self.final_category.keys())[0]
        max_cat_list = []
        for i in list(range(2,sheet1.max_row-1)):
            if self.name in sheet1.cell(i,2).value:
                if max_category in sheet1.cell(i,4).value:
                    max_cat_list.append(sheet1.cell(i,3).value)

        max_cat_str = ' '.join(max_cat_list)
        
        #알파벳을 제외한 문자 제거
        sen = re.sub("[^a-zA-Z]"," ",max_cat_str)
        low_sen = sen.lower()
        words = low_sen.split()

        remained_words = [w for w in words if not w in stopwords.words('english')]

        stemmer = SnowballStemmer('english')
        stem_words = [stemmer.stem(w) for w in remained_words]       
        
        word_text = ' '.join(stem_words)
        wordcloud = WordCloud(stopwords = STOPWORDS, 
                            background_color = 'white', width = 800, height = 600).generate(word_text)
        plt.figure(figsize = (15 , 10))
        plt.imshow(wordcloud)
        plt.axis("off")

        plt.savefig('{}_04_word.png'.format(self.name),dpi=300)
        b = '{}_04_word.png'.format(self.name)
        return b


    def WordSearch(self):
        max_category = list(self.final_category.keys())[0]
        max_cat_list = []
        for i in list(range(2,sheet1.max_row-1)):
            if self.name in sheet1.cell(i,2).value:
                if max_category in sheet1.cell(i,4).value:
                    max_cat_list.append(sheet1.cell(i,3).value)

        max_cat_str = ' '.join(max_cat_list)
        
        #알파벳을 제외한 문자 제거
        sen = re.sub("[^a-zA-Z]"," ",max_cat_str)
        low_sen = sen.lower()
        words = low_sen.split()

        #불용어 제거
        remained_words = [w for w in words if not w in stopwords.words('english')]

        #어간추출(Stemming)_정해진 규칙으로 단어의 어미를 자르는 작업_섬세한 작업X
        #stemmer = SnowballStemmer('english')
        #stem_words = [stemmer.stem(w) for w in remained_words]

        #표제어_기본 사전형 단어(Lemmatization)
        n=WordNetLemmatizer()
        lem = [n.lemmatize(w) for w in remained_words]


        #최다빈도 단어 추출
        text_num = {}        
        #for i in stem_words:
        for i in lem:
            if not i in text_num:
                text_num[i] = 1
            else:
                text_num[i] += 1

        def f1(x):
            return text_num[x]

        topfive_text = {}
        for i in range(5):
            key_max = max(text_num.keys(), key=f1)
            topfive_text[key_max] = text_num[key_max]
            del text_num[key_max]

        #유사도 검사(category-최대빈도 5단어)
        category_sim = wordnet.synsets(max_category)[0]

        word_sim_dict = {}
        for i in list(topfive_text.keys()):
            input_word = wordnet.synsets(i)[0]
            word_sim_dict[i] = round(input_word.wup_similarity(category_sim),3)

        #빈도수(0~1 범위) 비율로 변환
        sum_text =0
        for i in topfive_text.values():
            sum_text += int(i)
        
        #유사도와 빈도수의 가중치를 조정하여 최종 단어 결정
        final_word_cacul = {}
        for i in list(word_sim_dict.keys()):
            final_word_cacul[i] = 0.7*word_sim_dict[i] + 0.3*(topfive_text[i]/sum_text)

        def f1(x):
            return final_word_cacul[x]

        max_select_word = {}
        for i in range(5):
            word_max = max(final_word_cacul.keys(), key=f1)
            max_select_word[word_max] = final_word_cacul[word_max]
            del final_word_cacul[word_max]    

        select_word = max_category + ' '+ list(max_select_word.keys())[0]

        #return select_word

        driver = webdriver.Chrome(r"C:\Users\신충현\Downloads\chromedriver_win32\chromedriver.exe")
        driver.implicitly_wait(3)
        driver.get("http://www.yes24.com/24/Category/NewProduct")
        assert "YES24" in driver.title

        search = driver.find_element_by_xpath('//*[@id="query"]')
        search.send_keys(select_word)
        search.send_keys(Keys.ENTER)   
        
        #페이지 이동 후 BeautifulSoup 라이브러리 통해 html 인식
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')

        #페이지에서 원하는 정보(제목) 출력
        schwrap = soup.find("div",{"id":"schContent_wrap"})
        goodlist = schwrap.find("div",{"class":"goodsList goodsList_list"})
        infogrp = goodlist.find_all("td",{"class" : "goods_infogrp"})
        
        titles = []
        for i in infogrp:
            titles.append(i.find("strong").text)        
        title_three = titles[0:3]

        return "\n".join(title_three)

    def Book_img(self):
        max_category = list(self.final_category.keys())[0]
        max_cat_list = []
        for i in list(range(2,sheet1.max_row-1)):
            if self.name in sheet1.cell(i,2).value:
                if max_category in sheet1.cell(i,4).value:
                    max_cat_list.append(sheet1.cell(i,3).value)

        max_cat_str = ' '.join(max_cat_list)
        
        #알파벳을 제외한 문자 제거
        sen = re.sub("[^a-zA-Z]"," ",max_cat_str)
        low_sen = sen.lower()
        words = low_sen.split()

        #불용어 제거
        remained_words = [w for w in words if not w in stopwords.words('english')]

        #어간추출(Stemming)_정해진 규칙으로 단어의 어미를 자르는 작업_섬세한 작업X
        #stemmer = SnowballStemmer('english')
        #stem_words = [stemmer.stem(w) for w in remained_words]

        #표제어_기본 사전형 단어(Lemmatization)
        n=WordNetLemmatizer()
        lem = [n.lemmatize(w) for w in remained_words]


        #최다빈도 단어 추출
        text_num = {}        
        #for i in stem_words:
        for i in lem:
            if not i in text_num:
                text_num[i] = 1
            else:
                text_num[i] += 1

        def f1(x):
            return text_num[x]

        topfive_text = {}
        for i in range(5):
            key_max = max(text_num.keys(), key=f1)
            topfive_text[key_max] = text_num[key_max]
            del text_num[key_max]

        #유사도 검사(category-최대빈도 5단어)
        category_sim = wordnet.synsets(max_category)[0]

        word_sim_dict = {}
        for i in list(topfive_text.keys()):
            input_word = wordnet.synsets(i)[0]
            word_sim_dict[i] = round(input_word.wup_similarity(category_sim),3)

        #빈도수(0~1 범위) 비율로 변환
        sum_text =0
        for i in topfive_text.values():
            sum_text += int(i)
        
        #유사도와 빈도수의 가중치를 조정하여 최종 단어 결정
        final_word_cacul = {}
        for i in list(word_sim_dict.keys()):
            final_word_cacul[i] = 0.7*word_sim_dict[i] + 0.3*(topfive_text[i]/sum_text)

        def f1(x):
            return final_word_cacul[x]

        max_select_word = {}
        for i in range(5):
            word_max = max(final_word_cacul.keys(), key=f1)
            max_select_word[word_max] = final_word_cacul[word_max]
            del final_word_cacul[word_max]    

        select_word = max_category + ' '+ list(max_select_word.keys())[0]


        driver = webdriver.Chrome(r"C:\Users\신충현\Downloads\chromedriver_win32\chromedriver.exe")
        driver.implicitly_wait(3)
        driver.get("http://www.yes24.com/24/Category/NewProduct")
        assert "YES24" in driver.title

        search = driver.find_element_by_xpath('//*[@id="query"]')
        search.send_keys(select_word)
        search.send_keys(Keys.ENTER)   
        
        driver.implicitly_wait(3)
        driver.find_element_by_xpath('//*[@id="schMid_wrap"]/div[3]/div[2]/table/tbody/tr[1]/td[1]/a/img').click()
   
        #페이지 이동 후 BeautifulSoup 라이브러리 통해 html 인식
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')

        #페이지에서 원하는 정보(이미지 페이지)출력
        soup = soup.find("div",{"class":"gd_imgArea"})
        return soup.find("img")["src"]
           




