"""목표는 사람의 이름을 검색하면, 해당되는 사람에게 필요한 정보를 추천하는 것이다.
월말에 한 달 동안 업무 시간의 비율을 바탕으로 자신에게 필요한 신간 도서를 추천한다.
그리고 아직 진행 중에 있는 일에 대한 자료를 자동 검색을 통해 추천한다. 
추후에는 추천에 머무르는 것이 아닌 신뢰성과 필요에 맞는 적합성을 겸비하여
사용자로 하여금 업무에 도움이 되는 프로그램을 목표로 하고 있다.""" 

# %%
#엑셀 내용을 참조
import openpyxl
wb= openpyxl.load_workbook(r"C:\Users\신충현\Desktop\project\All_Task.xlsx")
sheet1 = wb.get_sheet_by_name("04_month")

#이름 입력
name = "ih".upper()
name_category = {}

#해당 이름에 해당하는 분야별 소요시간 합계 출력
for i in list(range(2,sheet1.max_row-1)):
    if name in sheet1.cell(i,2).value:
        if  sheet1.cell(i,4).value not in name_category:
            name_category[sheet1.cell(i,4).value] = float(sheet1.cell(i,5).value)
        else:
            name_category[sheet1.cell(i,4).value] += float(sheet1.cell(i,5).value)
    else:
        continue

#총 분야의 소요시간 합계
sum_value = 0
for i in name_category.values():
    sum_value += i

#시간을 percentage 단위로 변환
for key,element in name_category.items():
    name_category[key] = round((element/sum_value)*100,2)

#시간에 대해 최댓값을 갖는 3가지 분야만 도출
final_category = {}
for i in range(3):
    value_list = list(name_category.values())
    key_list = list(name_category.keys())
    max_key = key_list[value_list.index(max(value_list))]
    final_category[max_key] = max(value_list)
    del name_category[max_key]

#남은 부분 계산
three_sum = 0
for i in final_category.values():
    three_sum += i
residual = 100 - three_sum

#남은 부분 추가
final_category["residual"] = round(residual,2)

#입력된 사람의 분야별 소요시간(최대 3)
print(name, "->", ",".join("{}:{}".format(key,value) for key, value in final_category.items())) 
#%%
#해당 이름에 대한 workable 일의 Task Name 리스트 출력
list_task_name = []
for i in list(range(2,631)):
    if name in sheet1.cell(i,2).value:
        if  sheet1.cell(i,7).value == "Workable":
            list_task_name.append(sheet1.cell(i,3).value)
    else:
        continue

#사람마다 일의 비중을 토대로 자료 찾아주는 알고리즘(일의 비중 -> 추천 도서, workable -> 구글 자료)
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import time

# workable -> 구글 자료

#구글 검색창에 workable의 task name 입력
driver_google = webdriver.Chrome(r"C:\Users\신충현\Downloads\chromedriver_win32\chromedriver.exe")
driver_google.implicitly_wait(3)
driver_google.get("https://www.google.com/search?q={}".format(list_task_name[0]))

#검색 결과 페이지에서 상위 5개의 자료(url) 추천
results = driver_google.find_elements_by_css_selector('div.g')
hrefs = []
for i in range(5):
    link = results[i].find_element_by_tag_name("a")
    hrefs.append(link.get_attribute("href"))

for href in hrefs:
   print(href)

for href in hrefs:
   driver_google.execute_script('window.open("' + href + '","_blank");')



