
# %%
#엑셀 내용을 참조
import openpyxl
wb= openpyxl.load_workbook(r"C:\Users\신충현\Desktop\project\All_Task.xlsx")
sheet1 = wb.get_sheet_by_name("04_month")
#이름 입력
name = "hh".upper()
name_category = {}

#해당 이름에 해당하는 분야별 소요시간 합계 출력
#for i in list(range(2,631)):
for i in list(range(2,sheet1.max_row-1)):
    if name in sheet1.cell(i,2).value:
        if  sheet1.cell(i,4).value not in name_category:
            name_category[sheet1.cell(i,4).value] = float(sheet1.cell(i,5).value)
        else:
            name_category[sheet1.cell(i,4).value] += float(sheet1.cell(i,5).value)
    else:
        continue

#총 분야의 소요시간 합계
sum_value = 0
for i in name_category.values():
    sum_value += i

#시간을 percentage 단위로 변환
for key,element in name_category.items():
    name_category[key] = round((element/sum_value)*100,2)

#시간에 대해 최댓값을 갖는 3가지 분야만 도출
final_category = {}
for i in range(3):
    value_list = list(name_category.values())
    key_list = list(name_category.keys())
    max_key = key_list[value_list.index(max(value_list))]
    final_category[max_key] = max(value_list)
    del name_category[max_key]

#남은 부분 계산
three_sum = 0
for i in final_category.values():
    three_sum += i
residual = 100 - three_sum

#남은 부분 추가
final_category["residual"] = round(residual,2)

#입력된 사람의 분야별 소요시간(최대 3)
print(name, "->", ",".join("{}:{}".format(key,value) for key, value in final_category.items())) 

# %%
#가장 많은 비중을 차지하는 category의 task name을 모두 불러오기
max_category = list(final_category.keys())[0]
max_cat_list = []
for i in list(range(2,sheet1.max_row-1)):
    if name in sheet1.cell(i,2).value:
        if max_category in sheet1.cell(i,4).value:
            max_cat_list.append(sheet1.cell(i,3).value)

max_cat_str = ' '.join(max_cat_list)
#print(max_cat_str)

#알파벳을 제외한 문자 제거
import re
sen = re.sub("[^a-zA-Z]"," ",max_cat_str)
low_sen = sen.lower()
words = low_sen.split()
#print(len(words))
#print(words)

#불용어 제거 
import nltk
from nltk.corpus import stopwords
#print(stopwords.words('english')[:10])

remained_words = [w for w in words if not w in stopwords.words('english')]
#print(len(remained_words))
#print(remained_words)

#어간추출
from nltk.stem.snowball import SnowballStemmer
stemmer = SnowballStemmer('english')
stem_words = [stemmer.stem(w) for w in remained_words]

import matplotlib
#print(c_nltk_text.plot(50))

text_num = {}
for i in stem_words:
    if not i in text_num:
        text_num[i] = 1
    else:
        text_num[i] += 1

print(text_num)
#%%
#사람마다 일의 비중을 토대로 자료 찾아주는 알고리즘(일의 비중 -> 추천 도서, workable -> 구글 자료)
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import time

# 일의 비중 -> 추천 도서

#yes24 웹페이지 이동
driver = webdriver.Chrome(r"C:\Users\신충현\Downloads\chromedriver_win32\chromedriver.exe")
driver.implicitly_wait(3)
driver.get("http://www.yes24.com/24/Category/NewProduct")
assert "YES24" in driver.title

#검색 창에 category 비율 가장 높은 keyword 자동입력
search = driver.find_element_by_xpath('//*[@id="query"]')
search.send_keys(list(final_category.keys())[0])
search.send_keys(Keys.ENTER)

#페이지 이동 후 BeautifulSoup 라이브러리 통해 html 인식
req = driver.page_source
soup = BeautifulSoup(req, 'html.parser')

#페이지에서 원하는 정보(제목) 출력
schwrap = soup.find("div",{"id":"schContent_wrap"})
goodlist = schwrap.find("div",{"class":"goodsList goodsList_list"})
infogrp = goodlist.find_all("td",{"class" : "goods_infogrp"})
titles = []
for i in infogrp:
    titles.append(i.find("strong").text)
title_three = titles[0:3]
print(title_three)


